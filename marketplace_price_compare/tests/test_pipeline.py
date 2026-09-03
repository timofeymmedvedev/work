"""Тесты разбора названий, фильтра релевантности и сборки отчёта.

Запуск: python -m unittest discover -s tests -v
"""

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from mpc.aggregate import drop_outliers, pick_matches, summarize
from mpc.normalize import build_query, tokenize
from mpc.relevance import Reference
from mpc.sources.base import SourceError, parse_price
from mpc.sources.ozon import OzonSource, _describe, extract_offers
from mpc.sources.wildberries import _extract_price


class TestQuery(unittest.TestCase):
    def test_отрезает_комплектацию_после_плюса(self):
        q = build_query("Пульт Radiomaster TX12 MKII ELRS + АКБ 18650 3000mAh (2шт)")
        self.assertEqual(q, "пульт radiomaster tx12 mkii elrs")

    def test_обрывает_запрос_на_артикуле(self):
        q = build_query(
            "Ноутбук ASUS Vivobook 16 X1605VA-MB2103 (90NB10N2-M02KW0) "
            "Intel Core i7-13620H 16GB DDR4 1TB"
        )
        self.assertEqual(q, "ноутбук asus vivobook 16 x1605va-mb2103")

    def test_десятичная_запятая_не_рвёт_токен(self):
        self.assertIn("1.75", tokenize("Пластик PLA 1,75 мм"))

    def test_пустое_название(self):
        self.assertEqual(build_query(None), "")
        self.assertEqual(build_query(""), "")


class TestRelevance(unittest.TestCase):
    def assertMatch(self, ref_name, candidate, expected):
        ok, _ = Reference(ref_name).match(candidate)
        self.assertEqual(ok, expected, f"{ref_name!r} vs {candidate!r}")

    def test_принимает_ту_же_модель(self):
        self.assertMatch(
            "Пульт Radiomaster TX12 MKII ELRS + АКБ 18650 3000mAh (2шт)",
            "RadioMaster TX12 MKII ELRS EdgeTX / Аппаратура управления fpv",
            True,
        )

    def test_отсекает_соседнюю_модель(self):
        self.assertMatch(
            "Мотор Surpass Hobby bat S5322 380KV",
            "Surpass Hobby S5322 200KV motor",
            False,
        )

    def test_отсекает_аксессуар(self):
        self.assertMatch(
            "Пульт Radiomaster TX12 MKII ELRS",
            "Чехол для пульта Radiomaster TX12 MKII ELRS",
            False,
        )

    def test_отсекает_другой_ноутбук_того_же_бренда(self):
        self.assertMatch(
            "Ноутбук ASUS Vivobook 16 X1605VA-MB2103 (90NB10N2-M02KW0)",
            "Ноутбук ASUS Vivobook 15 X1504VA-BQ2X",
            False,
        )

    def test_название_без_артикула_требует_совпадения_бренда(self):
        self.assertMatch("TBS Tango 2 PRO V4", "TBS Tango 2 Pro V4 пульт FPV", True)
        self.assertMatch("TBS Tango 2 PRO V4", "TBS Tango 2 стики", False)

    def test_чисто_кириллическое_название(self):
        self.assertMatch(
            "Очиститель тормозов TT Экстра 650 мл, аэрозоль (CT06/02)",
            "Очиститель карбюратора ABRO 500 мл",
            False,
        )


class TestAggregate(unittest.TestCase):
    def setUp(self):
        self.ref = Reference("Мотор Surpass Hobby bat S5322 380KV")

    def _offer(self, price, title="Мотор Surpass Hobby BAT S5322 380KV"):
        return {"title": title, "price": price, "url": ""}

    def test_берёт_только_первые_n_релевантных(self):
        offers = [self._offer(1000 + i * 10) for i in range(10)]
        picked = pick_matches(self.ref, offers, top_n=5, scan_depth=30, outlier_ratio=4.0)
        self.assertEqual(len(picked), 5)

    def test_пропускает_нерелевантные_но_идёт_вглубь(self):
        offers = [self._offer(999, "Совсем другой товар") for _ in range(5)]
        offers += [self._offer(1000), self._offer(1100)]
        picked = pick_matches(self.ref, offers, top_n=5, scan_depth=30, outlier_ratio=4.0)
        self.assertEqual(len(picked), 2)

    def test_игнорирует_нулевые_цены(self):
        picked = pick_matches(
            self.ref, [self._offer(0), self._offer(None), self._offer(1000)],
            top_n=5, scan_depth=30, outlier_ratio=4.0,
        )
        self.assertEqual(len(picked), 1)

    def test_отсекает_ценовой_выброс(self):
        matched = [{"price": p} for p in (1000, 1050, 1100, 30)]
        kept = drop_outliers(matched, ratio=4.0)
        self.assertEqual([m["price"] for m in kept], [1000, 1050, 1100])

    def test_не_отсекает_когда_осталось_бы_меньше_двух(self):
        matched = [{"price": p} for p in (10, 1000, 5000)]
        self.assertEqual(len(drop_outliers(matched, ratio=1.1)), 3)

    def test_сводка_пустой_группы(self):
        self.assertEqual(summarize([]), {"avg": None, "count": 0, "min": None, "max": None})

    def test_сводка_считает_среднее(self):
        s = summarize([{"price": 100}, {"price": 200}])
        self.assertEqual((s["avg"], s["count"], s["min"], s["max"]), (150.0, 2, 100, 200))


class TestPriceParsing(unittest.TestCase):
    def test_разбор_строки_с_рублём(self):
        self.assertEqual(parse_price("25 250 ₽"), 25250.0)

    def test_разбор_числа(self):
        self.assertEqual(parse_price(1234.5), 1234.5)

    def test_мусор_даёт_none(self):
        self.assertIsNone(parse_price("нет в наличии"))
        self.assertIsNone(parse_price(None))

    def test_wb_новая_схема_цены(self):
        product = {"sizes": [{"price": {"total": 2525000, "product": 2600000}}]}
        self.assertEqual(_extract_price(product), 25250.0)

    def test_wb_старая_схема_цены(self):
        self.assertEqual(_extract_price({"salePriceU": 990000}), 9900.0)

    def test_wb_без_цены(self):
        self.assertIsNone(_extract_price({"sizes": [{"price": {}}]}))


class TestOzonExtraction(unittest.TestCase):
    def test_достаёт_товар_из_widget_states(self):
        import json
        state = {
            "items": [{
                "action": {"link": "/product/radiomaster-tx12-885980980/?x=1"},
                "mainState": [
                    {"atom": {"textAtom": {"text": "Аппаратура управления Radiomaster TX12 MKII ELRS"}}},
                    {"atom": {"priceV2": {"price": [
                        {"text": "25 250 ₽", "textStyle": "PRICE"},
                        {"text": "48 601 ₽", "textStyle": "ORIGINAL_PRICE"},
                    ]}}},
                ],
            }]
        }
        payload = {"widgetStates": {"searchResultsV2-1-default-1": json.dumps(state)}}
        offers = extract_offers(payload)
        self.assertEqual(len(offers), 1)
        self.assertEqual(offers[0]["price"], 25250.0)
        self.assertIn("Radiomaster TX12", offers[0]["title"])
        self.assertTrue(offers[0]["url"].startswith("https://www.ozon.ru/product/"))

    def test_пустой_ответ(self):
        self.assertEqual(extract_offers({}), [])
        self.assertEqual(extract_offers({"widgetStates": {"x": "не json"}}), [])


class FakePage:
    """Страница-заглушка: отдаёт заданный текст и считает переходы."""

    def __init__(self, browser, body):
        self.browser = browser
        self.body = body

    def goto(self, url, **kwargs):
        self.browser.visited.append(url)

    def inner_text(self, _selector):
        return self.body

    def wait_for_selector(self, *a, **kw):
        pass

    def wait_for_timeout(self, *a, **kw):
        pass

    def evaluate(self, _script):
        return self.browser.dom_items

    def close(self):
        pass


class FakeBrowser:
    def __init__(self, api_body, dom_items=None):
        self.api_body = api_body
        self.dom_items = dom_items or []
        self.visited = []

    def new_page(self):
        return FakePage(self, self.api_body)


class TestOzonFallback(unittest.TestCase):
    def test_прогревает_сессию_перед_первым_запросом(self):
        import json
        payload = {"widgetStates": {"w": json.dumps({
            "items": [{
                "action": {"link": "/product/x-1/"},
                "mainState": [
                    {"atom": {"textAtom": {"text": "Мотор Surpass Hobby BAT S5322 380KV"}}},
                    {"atom": {"priceV2": {"price": [{"text": "5 000 ₽", "textStyle": "PRICE"}]}}},
                ],
            }]
        })}}
        browser = FakeBrowser(json.dumps(payload))
        offers = OzonSource(browser).search("мотор s5322")
        self.assertTrue(browser.visited[0].startswith("https://www.ozon.ru/"))
        self.assertEqual(len(offers), 1)

    def test_переходит_на_страницу_поиска_когда_api_не_json(self):
        browser = FakeBrowser(
            "<html>обычная страница</html>",
            dom_items=[{"title": "Мотор Surpass Hobby BAT S5322 380KV",
                        "price": "5000", "url": "/product/x-1/"}],
        )
        offers = OzonSource(browser).search("мотор s5322")
        self.assertEqual(len(offers), 1)
        self.assertEqual(offers[0]["price"], 5000.0)
        self.assertTrue(any("/search/" in u for u in browser.visited))

    def test_обе_попытки_провалились(self):
        browser = FakeBrowser("Доступ ограничен: проверка")
        with self.assertRaises(SourceError):
            OzonSource(browser).search("мотор s5322")

    def test_описание_антибот_страницы(self):
        self.assertEqual(_describe("Доступ ограничен"), "страница антибот-проверки")
        self.assertEqual(_describe(""), "пустой ответ")
        self.assertIn("обычный текст", _describe("обычный текст ответа"))


class TestExcelReport(unittest.TestCase):
    def test_дописывает_столбцы_и_считает_отклонение(self):
        import openpyxl
        from mpc.excel_io import read_products, write_report

        with tempfile.TemporaryDirectory() as tmp:
            src = os.path.join(tmp, "in.xlsx")
            out = os.path.join(tmp, "out.xlsx")
            book = openpyxl.Workbook()
            ws = book.active
            ws.append(["Название", "Категория", "Цена"])
            ws.append(["Мотор Surpass Hobby bat S5322 380KV", "Двигатели", 5717.89])
            ws.append(["Товар без выдачи", "Прочее", 100])
            book.save(src)

            products = read_products(src)
            self.assertEqual(len(products), 2)

            results = {
                (2, "ozon"): {"avg": 6000.0, "count": 5},
                (2, "yandex"): {"avg": 6200.0, "count": 3},
                (2, "wb"): {"avg": 5800.0, "count": 5},
            }
            write_report(src, out, products, results, ["ozon", "yandex", "wb"])

            check = openpyxl.load_workbook(out)
            ws2 = check.active
            headers = [c.value for c in ws2[1]]
            self.assertIn("Ozon: средняя цена", headers)
            self.assertIn("Средняя по маркетплейсам", headers)

            market_col = headers.index("Средняя по маркетплейсам") + 1
            delta_col = headers.index("Отклонение от рынка") + 1
            self.assertEqual(ws2.cell(row=2, column=market_col).value, 6000.0)
            # 5717.89 против рынка 6000 — наша цена ниже почти на 5 %.
            self.assertAlmostEqual(ws2.cell(row=2, column=delta_col).value, -0.047, places=3)
            # По второй строке выдачи не было — ячейки остаются пустыми.
            self.assertIsNone(ws2.cell(row=3, column=market_col).value)


if __name__ == "__main__":
    unittest.main()
