"""Чтение исходного прайса и сборка итогового файла."""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SOURCE_TITLES = {
    "ozon": "Ozon",
    "yandex": "Яндекс Маркет",
    "wb": "Wildberries",
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FORMAT = "# ##0.00 ₽"
PERCENT_FORMAT = "0.0%"


def read_products(path, sheet=None):
    """Возвращает [{row, name, category, price}] в порядке строк файла."""
    book = openpyxl.load_workbook(path, data_only=True)
    ws = book[sheet] if sheet else book.worksheets[0]

    products = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[0] if len(row) > 0 else None
        if not name:
            continue
        products.append({
            "row": idx,
            "name": str(name).strip(),
            "category": row[1] if len(row) > 1 else None,
            "price": row[2] if len(row) > 2 else None,
        })
    book.close()
    return products


def write_report(src_path, out_path, products, results, sources, sheet=None):
    """Дописывает к исходному листу столбцы средних цен по площадкам.

    Исходный файл открывается и пересохраняется целиком, чтобы сохранить
    порядок строк и всё, что в нём уже было: результат — тот же прайс с
    добавленными колонками, а не отдельная таблица.
    """
    book = openpyxl.load_workbook(src_path)
    ws = book[sheet] if sheet else book.worksheets[0]

    start = ws.max_column + 1
    columns = _build_columns(sources)

    for offset, (title, _) in enumerate(columns):
        cell = ws.cell(row=1, column=start + offset, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    by_row = {p["row"]: p for p in products}

    for row_idx in range(2, ws.max_row + 1):
        product = by_row.get(row_idx)
        if not product:
            continue
        values = _row_values(product, results, sources)
        for offset, (title, key) in enumerate(columns):
            cell = ws.cell(row=row_idx, column=start + offset, value=values.get(key))
            if key.endswith("_avg") or key == "market_avg":
                cell.number_format = MONEY_FORMAT
            elif key == "delta":
                cell.number_format = PERCENT_FORMAT

    _autosize(ws, start, start + len(columns) - 1)
    ws.freeze_panes = "A2"
    book.save(out_path)
    book.close()


def _build_columns(sources):
    columns = []
    for src in sources:
        label = SOURCE_TITLES.get(src, src)
        columns.append((f"{label}: средняя цена", f"{src}_avg"))
        columns.append((f"{label}: позиций", f"{src}_count"))
    columns.append(("Средняя по маркетплейсам", "market_avg"))
    columns.append(("Отклонение от рынка", "delta"))
    return columns


def _row_values(product, results, sources):
    values = {}
    averages = []
    for src in sources:
        entry = results.get((product["row"], src)) or {}
        avg = entry.get("avg")
        values[f"{src}_avg"] = avg
        values[f"{src}_count"] = entry.get("count") or 0
        if avg:
            averages.append(avg)

    market = round(sum(averages) / len(averages), 2) if averages else None
    values["market_avg"] = market

    own = product.get("price")
    if market and isinstance(own, (int, float)) and own:
        # Положительное значение — наша цена выше рынка.
        values["delta"] = round((own - market) / market, 4)
    else:
        values["delta"] = None
    return values


def _autosize(ws, first, last):
    for col in range(first, last + 1):
        letter = get_column_letter(col)
        header = ws.cell(row=1, column=col).value or ""
        ws.column_dimensions[letter].width = min(max(14, len(str(header)) * 0.9), 26)
